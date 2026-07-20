"""
Test XLSXExporter — testování strukturovaného exportu do Excel formátu.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import numpy as np

from core.peak import Peak
from core.project import Project
from core.spectrum import Spectrum
from file_io.xlsx_exporter import XLSXExporter


class TestXLSXExporter:
    """Test XLSX export functionality."""

    def test_export_peaks_only(self) -> None:
        """Export without a spectrum still produces a Peaks sheet with lean columns."""
        peaks = [
            Peak(
                position=1650.5, intensity=0.85, vibration_ids=[1], vibration_labels=["C=O stretch"]
            ),
            Peak(
                position=2950.0, intensity=0.92, vibration_ids=[2], vibration_labels=["C-H stretch"]
            ),
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            XLSXExporter().export(peaks, tmp_path)

            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(tmp_path)
            assert wb.sheetnames == ["Metadata", "Peaks"]

            peaks_ws = wb["Peaks"]
            assert peaks_ws.cell(1, 1).value == "Position (cm⁻¹)"
            assert peaks_ws.cell(1, 3).value == "Rel. intensity"
            assert peaks_ws.cell(1, 4).value == "Assignment"
            assert peaks_ws.cell(1, 5).value is None  # no extra "Assignments" count column

            assert peaks_ws.cell(2, 1).value == 2950
            assert peaks_ws.cell(2, 2).value == 0.92
            assert peaks_ws.cell(2, 3).value == "vs"
            assert peaks_ws.cell(2, 4).value == "C-H stretch"
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_export_with_spectrum_has_three_sheets_and_no_chart(self) -> None:
        """Export with a spectrum produces Metadata + Peaks + Spectrum sheets, no chart."""
        wavenumbers = np.linspace(4000, 400, 100)
        intensities = np.random.random(100)
        spectrum = Spectrum(wavenumbers=wavenumbers, intensities=intensities, title="Test Spectrum")
        project = Project(name="S-1", spectrum=spectrum)
        project.metadata.operator = "Analyst"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            XLSXExporter().export(
                [Peak(position=1650.5, intensity=0.85)], tmp_path, spectrum, project=project
            )

            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(tmp_path)
            assert wb.sheetnames == ["Metadata", "Peaks", "Spectrum"]

            meta_ws = wb["Metadata"]
            meta_values = {
                meta_ws.cell(r, 1).value: meta_ws.cell(r, 2).value
                for r in range(1, meta_ws.max_row + 1)
            }
            assert meta_values.get("Operator") == "Analyst"
            assert "Data points" not in meta_values  # trimmed to PDF-header fields

            spectrum_ws = wb["Spectrum"]
            assert spectrum_ws.cell(1, 1).value == "Wavenumber (cm⁻¹)"
            assert spectrum_ws.cell(2, 1).value == 4000.0
            assert isinstance(spectrum_ws.cell(2, 2).value, float)
            assert len(spectrum_ws._charts) == 0
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_export_empty_peaks(self) -> None:
        """Export with no peaks still writes headers and no data rows."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            XLSXExporter().export([], tmp_path)

            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(tmp_path)
            peaks_ws = wb["Peaks"]
            assert peaks_ws.cell(1, 1).value == "Position (cm⁻¹)"
            assert peaks_ws.cell(2, 1).value is None
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_export_peak_table_omits_unassigned_by_default(self) -> None:
        """Peaks sheet keeps PDF-consistent ordering and only assigned peaks by default."""
        peaks = [
            Peak(position=1712.6, intensity=11.2, label="1713"),
            Peak(position=2954.8, intensity=5.1, vibration_ids=[1], vibration_labels=["ν(C-H)"]),
            Peak(position=1456.4, intensity=3.7, vibration_ids=[2], vibration_labels=["δ(CH₂)"]),
        ]
        spectrum = Spectrum(
            wavenumbers=np.array([4000.0, 3000.0]), intensities=np.array([1.0, 2.0])
        )

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            XLSXExporter().export(peaks, tmp_path, spectrum)

            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(tmp_path)
            peaks_ws = wb["Peaks"]

            assert peaks_ws.cell(2, 1).value == 2955
            assert peaks_ws.cell(2, 4).value == "ν(C-H)"
            assert peaks_ws.cell(3, 1).value == 1456
            assert peaks_ws.cell(3, 4).value == "δ(CH₂)"
            assert peaks_ws.cell(4, 1).value is None  # unassigned peak omitted
        finally:
            tmp_path.unlink(missing_ok=True)
