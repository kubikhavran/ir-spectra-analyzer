"""
XLSXExporter — Export analysis to a structured Microsoft Excel workbook.

Zodpovědnost:
- Tři listy: Metadata, Peaks, Spectrum (plná X/Y data)
- Formátování buněk (tučné záhlaví, zamrzlé řádky, šířky sloupců)
- Vložený spojnicový graf spektra, aby si zákazník rovnou viděl křivku

Závislost: openpyxl
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from core.peak import Peak
from core.peak_assignments import build_peak_assignment_rows, collect_export_metadata
from core.spectrum import Spectrum
from file_io.atomic_output import atomic_output_path
from file_io.spreadsheet_safety import neutralize_spreadsheet_formula

if TYPE_CHECKING:
    from core.project import Project


class XLSXExporter:
    """Exports the full analysis to a formatted Excel (.xlsx) workbook."""

    def export(
        self,
        peaks: list[Peak],
        output_path: Path,
        spectrum: Spectrum | None = None,
        include_unassigned: bool = False,
        project: Project | None = None,
    ) -> None:
        """Export the analysis to an xlsx workbook.

        Args:
            peaks: List of peaks to export.
            output_path: Destination .xlsx file path.
            spectrum: Spectrum providing X/Y data and dip/absorption polarity.
            include_unassigned: Whether to include peaks without vibration assignments.
            project: Optional project for editable metadata (sample, operator…).
        """
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="374A6B")
        title_font = Font(bold=True, size=13)
        y_unit_label = spectrum.y_unit.value if spectrum is not None else "Intensity"

        def _style_header(ws, row: int, n_cols: int) -> None:
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        def _autosize(ws, max_width: int = 60) -> None:
            for col in ws.columns:
                length = 0
                letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        length = max(length, len(str(cell.value)))
                ws.column_dimensions[letter].width = min(length + 2, max_width)

        wb = openpyxl.Workbook()

        # ── Metadata sheet ──────────────────────────────────────────────────
        meta_ws = wb.active
        meta_ws.title = "Metadata"
        meta_ws.cell(row=1, column=1, value="IR spectrum analysis").font = title_font
        meta_ws.cell(row=3, column=1, value="Field")
        meta_ws.cell(row=3, column=2, value="Value")
        _style_header(meta_ws, 3, 2)
        for offset, (label, value) in enumerate(
            collect_export_metadata(project, spectrum), start=4
        ):
            meta_ws.cell(row=offset, column=1, value=label).font = Font(bold=True)
            meta_ws.cell(
                row=offset,
                column=2,
                value=neutralize_spreadsheet_formula(value),
            )
        _autosize(meta_ws)

        # ── Peaks sheet (assigned peaks unless explicitly asked otherwise) ──
        peaks_ws = wb.create_sheet("Peaks")
        peaks_headers = [
            "Position (cm⁻¹)",
            f"Intensity ({y_unit_label})",
            "Rel. intensity",
            "Assignment",
        ]
        for col, header in enumerate(peaks_headers, start=1):
            peaks_ws.cell(row=1, column=col, value=header)
        _style_header(peaks_ws, 1, len(peaks_headers))

        assignment_rows = build_peak_assignment_rows(
            peaks,
            is_dip_spectrum=spectrum.is_dip_spectrum if spectrum is not None else False,
            include_unassigned=include_unassigned,
        )
        for row, assignment_row in enumerate(assignment_rows, start=2):
            peaks_ws.cell(row=row, column=1, value=assignment_row.position)
            peaks_ws.cell(row=row, column=2, value=round(assignment_row.intensity, 4))
            peaks_ws.cell(row=row, column=3, value=assignment_row.intensity_label)
            peaks_ws.cell(
                row=row,
                column=4,
                value=neutralize_spreadsheet_formula(assignment_row.assignment),
            )
        peaks_ws.freeze_panes = "A2"
        _autosize(peaks_ws)

        # ── Spectrum sheet (full X/Y data, plottable) ───────────────────────
        if spectrum is not None:
            self._write_spectrum_sheet(wb, spectrum, y_unit_label)

        with atomic_output_path(output_path) as temp_path:
            wb.save(temp_path)

    @staticmethod
    def _write_spectrum_sheet(wb, spectrum: Spectrum, y_unit_label: str) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415

        ws = wb.create_sheet("Spectrum")
        ws.cell(row=1, column=1, value="Wavenumber (cm⁻¹)")
        ws.cell(row=1, column=2, value=f"Intensity ({y_unit_label})")
        for col in (1, 2):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="374A6B")
            cell.alignment = Alignment(horizontal="center")

        for row, (wn, intensity) in enumerate(
            zip(spectrum.wavenumbers, spectrum.intensities, strict=True), start=2
        ):
            ws.cell(row=row, column=1, value=round(float(wn), 2))
            ws.cell(row=row, column=2, value=round(float(intensity), 6))

        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 22
