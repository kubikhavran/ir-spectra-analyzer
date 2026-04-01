"""
XLSXExporter — Export peakové tabulky do Microsoft Excel formátu.

Zodpovědnost:
- Export peaků do formátovaného .xlsx souboru
- Formátování buněk (tučné záhlaví, zarovnání čísel)
- Metadata spektra v prvních řádcích

Závislost: openpyxl
"""
from __future__ import annotations

from pathlib import Path

from core.peak import Peak
from core.spectrum import Spectrum


class XLSXExporter:
    """Exports peak table to a formatted Excel (.xlsx) workbook."""

    def export(
        self,
        peaks: list[Peak],
        output_path: Path,
        spectrum: Spectrum | None = None,
    ) -> None:
        """Export peaks to an xlsx workbook.

        Args:
            peaks: List of peaks to export.
            output_path: Destination .xlsx file path.
            spectrum: Optional spectrum for metadata rows.
        """
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import Font  # noqa: PLC0415

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IR Peaks"  # type: ignore[union-attr]

        headers = ["Position (cm⁻¹)", "Intensity", "Label", "Vibration", "Notes"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)  # type: ignore[union-attr]
            cell.font = Font(bold=True)

        for row, peak in enumerate(peaks, start=2):
            ws.cell(row=row, column=1, value=round(peak.position, 2))  # type: ignore[union-attr]
            ws.cell(row=row, column=2, value=round(peak.intensity, 4))  # type: ignore[union-attr]
            ws.cell(row=row, column=3, value=peak.label)  # type: ignore[union-attr]

        wb.save(output_path)
